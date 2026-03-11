"""
file_handler.py
===============
File read/write utilities for CSV, JSON, JSONL, and plain-text files.

Uses only the Python standard library.  Optional Excel support is
available when ``openpyxl`` is installed.

Usage::

    from wingbrace.utilities import FileHandler

    # Read CSV into list of dicts
    records = FileHandler.read_csv("data/users.csv")

    # Write list of dicts to JSON
    FileHandler.write_json(records, "output/users.json")

    # Append to a CSV
    FileHandler.append_csv([{"id": 5, "name": "Eve"}], "data/users.csv")

    # Read Excel (requires openpyxl)
    rows = FileHandler.read_excel("report.xlsx", sheet_name="Sheet1")
"""

import csv
import io
import json
import os
from typing import Any, Iterator


class FileHandler:
    """
    Collection of static helper methods for common file I/O operations.

    All methods are ``@staticmethod`` so the class can be used without
    instantiation, or you can create an instance if you prefer.
    """

    # ------------------------------------------------------------------
    # CSV
    # ------------------------------------------------------------------

    @staticmethod
    def read_csv(
        filepath: str,
        encoding: str = "utf-8",
        delimiter: str = ",",
        skip_blank_lines: bool = True,
    ) -> list[dict[str, str]]:
        """
        Read a CSV file into a list of dicts.

        Parameters
        ----------
        filepath:
            Path to the CSV file.
        encoding:
            File encoding.
        delimiter:
            Field delimiter character.
        skip_blank_lines:
            Skip rows where all fields are empty.

        Returns
        -------
        list of dict
            One dict per data row, keyed by header names.
        """
        with open(filepath, newline="", encoding=encoding) as fh:
            reader = csv.DictReader(fh, delimiter=delimiter)
            rows = list(reader)
        if skip_blank_lines:
            rows = [r for r in rows if any(v.strip() for v in r.values())]
        return rows

    @staticmethod
    def write_csv(
        data: list[dict[str, Any]],
        filepath: str,
        fieldnames: list[str] | None = None,
        encoding: str = "utf-8",
        delimiter: str = ",",
    ) -> None:
        """
        Write a list of dicts to a CSV file.

        Parameters
        ----------
        data:
            Rows to write.
        filepath:
            Destination file path.  Parent directories are created.
        fieldnames:
            Column order.  Defaults to keys of the first row.
        """
        if not data:
            return
        os.makedirs(os.path.dirname(os.path.abspath(filepath)), exist_ok=True)
        fieldnames = fieldnames or list(data[0].keys())
        with open(filepath, "w", newline="", encoding=encoding) as fh:
            writer = csv.DictWriter(
                fh, fieldnames=fieldnames, delimiter=delimiter, extrasaction="ignore"
            )
            writer.writeheader()
            writer.writerows(data)

    @staticmethod
    def append_csv(
        data: list[dict[str, Any]],
        filepath: str,
        encoding: str = "utf-8",
        delimiter: str = ",",
    ) -> None:
        """
        Append rows to an existing CSV file (creates it if absent).

        If the file already exists the header row is *not* written again.
        """
        if not data:
            return
        file_exists = os.path.exists(filepath)
        os.makedirs(os.path.dirname(os.path.abspath(filepath)), exist_ok=True)
        fieldnames = list(data[0].keys())
        with open(filepath, "a", newline="", encoding=encoding) as fh:
            writer = csv.DictWriter(
                fh, fieldnames=fieldnames, delimiter=delimiter, extrasaction="ignore"
            )
            if not file_exists:
                writer.writeheader()
            writer.writerows(data)

    @staticmethod
    def csv_to_string(
        data: list[dict[str, Any]],
        fieldnames: list[str] | None = None,
        delimiter: str = ",",
    ) -> str:
        """Return CSV-formatted string from *data* (no file I/O)."""
        if not data:
            return ""
        buf = io.StringIO()
        fieldnames = fieldnames or list(data[0].keys())
        writer = csv.DictWriter(buf, fieldnames=fieldnames, delimiter=delimiter)
        writer.writeheader()
        writer.writerows(data)
        return buf.getvalue()

    # ------------------------------------------------------------------
    # JSON
    # ------------------------------------------------------------------

    @staticmethod
    def read_json(filepath: str, encoding: str = "utf-8") -> Any:
        """
        Read and parse a JSON file.

        Returns
        -------
        Any
            The parsed JSON content (typically a dict or list).
        """
        with open(filepath, encoding=encoding) as fh:
            return json.load(fh)

    @staticmethod
    def write_json(
        data: Any,
        filepath: str,
        indent: int = 2,
        encoding: str = "utf-8",
        ensure_ascii: bool = False,
    ) -> None:
        """Write *data* to a JSON file."""
        os.makedirs(os.path.dirname(os.path.abspath(filepath)), exist_ok=True)
        with open(filepath, "w", encoding=encoding) as fh:
            json.dump(data, fh, indent=indent, ensure_ascii=ensure_ascii, default=str)

    # ------------------------------------------------------------------
    # JSON Lines (JSONL)
    # ------------------------------------------------------------------

    @staticmethod
    def read_jsonl(
        filepath: str, encoding: str = "utf-8"
    ) -> list[Any]:
        """Read a JSONL (newline-delimited JSON) file into a list."""
        records = []
        with open(filepath, encoding=encoding) as fh:
            for line in fh:
                line = line.strip()
                if line:
                    records.append(json.loads(line))
        return records

    @staticmethod
    def write_jsonl(
        data: list[Any],
        filepath: str,
        encoding: str = "utf-8",
        append: bool = False,
    ) -> None:
        """Write a list of objects to a JSONL file."""
        os.makedirs(os.path.dirname(os.path.abspath(filepath)), exist_ok=True)
        mode = "a" if append else "w"
        with open(filepath, mode, encoding=encoding) as fh:
            for record in data:
                fh.write(json.dumps(record, default=str) + "\n")

    # ------------------------------------------------------------------
    # Plain text
    # ------------------------------------------------------------------

    @staticmethod
    def read_text(filepath: str, encoding: str = "utf-8") -> str:
        """Read a text file and return its contents as a string."""
        with open(filepath, encoding=encoding) as fh:
            return fh.read()

    @staticmethod
    def write_text(
        content: str,
        filepath: str,
        encoding: str = "utf-8",
        append: bool = False,
    ) -> None:
        """Write *content* to a text file."""
        os.makedirs(os.path.dirname(os.path.abspath(filepath)), exist_ok=True)
        mode = "a" if append else "w"
        with open(filepath, mode, encoding=encoding) as fh:
            fh.write(content)

    @staticmethod
    def read_lines(
        filepath: str,
        encoding: str = "utf-8",
        strip: bool = True,
        skip_blank: bool = True,
    ) -> list[str]:
        """Read a text file and return a list of lines."""
        with open(filepath, encoding=encoding) as fh:
            lines = fh.readlines()
        if strip:
            lines = [ln.strip() for ln in lines]
        if skip_blank:
            lines = [ln for ln in lines if ln]
        return lines

    # ------------------------------------------------------------------
    # Excel (optional — requires openpyxl)
    # ------------------------------------------------------------------

    @staticmethod
    def read_excel(
        filepath: str,
        sheet_name: str | None = None,
        header_row: int = 1,
    ) -> list[dict[str, Any]]:
        """
        Read an Excel file into a list of dicts.

        Requires ``openpyxl`` (``pip install openpyxl``).

        Parameters
        ----------
        filepath:
            Path to the ``.xlsx`` file.
        sheet_name:
            Sheet to read.  Defaults to the first sheet.
        header_row:
            Row number (1-based) of the header.
        """
        try:
            import openpyxl
        except ImportError as exc:
            raise ImportError(
                "openpyxl is required to read Excel files: pip install openpyxl"
            ) from exc

        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(c) if c is not None else f"col_{i}"
                   for i, c in enumerate(rows[header_row - 1])]
        return [
            dict(zip(headers, row))
            for row in rows[header_row:]
        ]

    @staticmethod
    def write_excel(
        data: list[dict[str, Any]],
        filepath: str,
        sheet_name: str = "Sheet1",
        fieldnames: list[str] | None = None,
    ) -> None:
        """
        Write a list of dicts to an Excel ``.xlsx`` file.

        Requires ``openpyxl`` (``pip install openpyxl``).
        """
        try:
            import openpyxl
        except ImportError as exc:
            raise ImportError(
                "openpyxl is required to write Excel files: pip install openpyxl"
            ) from exc

        if not data:
            return
        os.makedirs(os.path.dirname(os.path.abspath(filepath)), exist_ok=True)
        fieldnames = fieldnames or list(data[0].keys())
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(fieldnames)
        for row in data:
            ws.append([row.get(f) for f in fieldnames])
        wb.save(filepath)

    # ------------------------------------------------------------------
    # Utility helpers
    # ------------------------------------------------------------------

    @staticmethod
    def ensure_dir(path: str) -> str:
        """
        Create *path* as a directory (including parents) if it does not
        exist.

        Returns
        -------
        str
            The resolved absolute path.
        """
        abs_path = os.path.abspath(path)
        os.makedirs(abs_path, exist_ok=True)
        return abs_path

    @staticmethod
    def file_exists(filepath: str) -> bool:
        """Return ``True`` if *filepath* exists and is a file."""
        return os.path.isfile(filepath)

    @staticmethod
    def list_files(
        directory: str,
        extension: str | None = None,
        recursive: bool = False,
    ) -> list[str]:
        """
        List files in *directory*, optionally filtered by *extension*.

        Parameters
        ----------
        extension:
            File extension filter, e.g. ``".csv"`` (include the dot).
        recursive:
            When ``True``, walk subdirectories.

        Returns
        -------
        list of str
            Absolute file paths.
        """
        results: list[str] = []
        if recursive:
            for root, _, files in os.walk(directory):
                for fname in files:
                    if extension is None or fname.endswith(extension):
                        results.append(os.path.join(root, fname))
        else:
            for entry in os.listdir(directory):
                full = os.path.join(directory, entry)
                if os.path.isfile(full) and (
                    extension is None or entry.endswith(extension)
                ):
                    results.append(full)
        return sorted(results)
